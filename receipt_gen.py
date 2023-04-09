import sys
import os
import re
import requests
from configparser import ConfigParser
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from random import randint
from urllib.parse import quote

#Google API key
api_key = "API_KEY"

def read_config():
    config = ConfigParser()
    config.read("config.ini")
    receipt_template = config.get("Template", "receipt")
    return receipt_template.strip()
def read_xlsx(source_file):
    wb = load_workbook(filename=source_file)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    return data

def generate_random_time():
    hour = randint(7, 18)
    minute = randint(0, 59)
    return f"{hour:02d}:{minute:02d}"

def generate_random_reference():
    reference = ""
    for i in range(7):
        if i == 0 or i == 1:
            reference += chr(randint(65, 90))
        else:
            reference += str(randint(0, 9))
    return reference

def generate_random_number(length):
    number = ""
    for i in range(length):
        number += str(randint(0, 9))
    return number

def generate_transaction_id(date):
    year = date[-2:]
    month = date[:2]
    day = date[3:5]
    return f"{year}{month}{day}{generate_random_number(10)}"

def get_business_info(name, desc):
    # Returns either name if it's not empty otherwise desc
    # if both are empty, returns "Business Name"
    if desc:
        return desc
    elif name:
        return name
    else:
        return "Business Name"

def get_card_info(card):
    return card[-5:-1]


def get_address(name, desc):
    if len(name) > len(desc):
        place_name = name
    else:
        place_name = desc
    place_name_search = quote(place_name + "around CT/NY")
    print("Parsing address for: " + place_name + " ...")
    base_url = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json?"
    input_type = "textquery"
    fields = "formatted_address"
    print("Querying Google API...")
    query_url = f"{base_url}input={place_name_search}&inputtype={input_type}&fields={fields}&key={api_key}"
    print(query_url)
    response = requests.get(query_url)
    if response.status_code == 200:
        data = response.json()
        if data.get('status') == 'OK':
            print("Found address: " + data['candidates'][0]['formatted_address']+ " ...\n")
            return str(data['candidates'][0]['formatted_address'])
        else:
            print("No address found for: " + place_name + " ...\n")
            return "Purchase Receipt"
    else:
        return f"Error: {response.status_code}"

def get_formatted_address(name, desc):
    address = get_address(name, desc)
    address = address.replace(", ", ",\n")
    return address

def total_no_tax(total):
    # Take in value then removes 6.35% tax
    # return with 2 decimal places
    return round(total / 1.0635, 2)

def format_data(data, template):
    formatted_data = []
    for row in data:
        if row[1] is not None and row[2] == "Expense":
            formatted_row = template.replace("[CATEGORY]", str(row[0]))
            formatted_row = formatted_row.replace("[DATE]", str(row[1]))
            formatted_row = formatted_row.replace("[TYPE]", str(row[2]))
            formatted_row = formatted_row.replace("[NUM]", str(row[3]))
            formatted_row = formatted_row.replace("[NAME]", str(row[5]))
            formatted_row = formatted_row.replace("[DESC]", str(row[6]))
            formatted_row = formatted_row.replace("[ACCOUNT]", str(row[7]))
            formatted_row = formatted_row.replace("[ACC]", str(row[8]))
            # Assigns string of total no tax amount rounded with 2 decimal places
            formatted_row = formatted_row.replace("[AMOUNT]", "{:.2f}".format(total_no_tax(row[9]), 2))
            formatted_row = formatted_row.replace("[TAX]", "{:.2f}".format(row[9] - total_no_tax(row[9]), 2))
            formatted_row = formatted_row.replace("[TOTAL]", "{:.2f}".format(row[9], 2))
            formatted_row = formatted_row.replace("[TIME]", generate_random_time())
            formatted_row = formatted_row.replace("[REF]", generate_random_reference())
            formatted_row = formatted_row.replace("[THREE]", generate_random_number(3))
            formatted_row = formatted_row.replace("[FOUR]", generate_random_number(4))
            formatted_row = formatted_row.replace("[FIVE]", generate_random_number(5))
            formatted_row = formatted_row.replace("[TRANS]", generate_transaction_id(str(row[1])))
            formatted_row = formatted_row.replace("[BUSINESS]", get_business_info(row[5], row[6]))
            formatted_row = formatted_row.replace("[CARD]", get_card_info(row[8]))
            formatted_row = formatted_row.replace("[ADDRESS]", get_formatted_address(str(row[5]), str(row[6])))
            formatted_data.append(formatted_row)
    return formatted_data

def create_receipt_image(text, output_folder, index):
    img = Image.new("RGB", (200, 300), color=(255, 255, 255))
    d = ImageDraw.Draw(img)
    font = ImageFont.truetype("MerchantCopy-GOXq.ttf", 14)
    d.text((10, 10), text, fill=(0, 0, 0), font=font)
    img.save(os.path.join(output_folder, f"receipt_{index}.png"))

def main(source_file, destination_folder=""):
    if not destination_folder:
        destination_folder = os.getcwd()

    template = read_config()
    data = read_xlsx(source_file)
    formatted_data = format_data(data, template)

    for index, receipt_text in enumerate(formatted_data):
        create_receipt_image(receipt_text, destination_folder, index)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python receipt_generator.py <source_file> [<destination_folder>]")
        sys.exit(1)

    source_file = sys.argv[1]
    destination_folder = sys.argv[2] if len(sys.argv) > 2 else ""
    main(source_file, destination_folder)
